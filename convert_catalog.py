from pathlib import Path
import json
import re
from openpyxl import load_workbook

EXCEL_FILE = Path("catalog.xlsx")
OUTPUT_FILE = Path("catalog-data.json")

# Leave as None to automatically find the worksheet containing the catalog.
# Or set a specific tab name, e.g. SHEET_NAME = "Xcvrs only"
SHEET_NAME = None

# How many rows at the top of each worksheet to scan for headers.
HEADER_SCAN_ROWS = 100

PART_NUMBER_HEADERS = {
    "part number",
    "t1nexus part number",
    "t1 pn",
    "pn",
    "part no",
    "part no.",
    "sku",
    "item number",
}

PRODUCT_DESCRIPTION_HEADERS = {
    "product description",
    "description",
    "product desc",
    "product desc.",
    "item description",
}

DATA_RATE_HEADERS = {
    "data rate",
    "datarate",
    "rate",
    "speed",
}

FORM_FACTOR_HEADERS = {
    "form factor",
    "form factor name",
    "formfactor",
    "form",
    "ff",
}


def normalize_header(value):
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value)
        .strip()
        .lower()
        .replace("_", " ")
        .replace("-", " "),
    )


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def find_column(headers, accepted_headers):
    accepted = {normalize_header(x) for x in accepted_headers}

    for column_number, header in enumerate(headers, start=1):
        if normalize_header(header) in accepted:
            return column_number

    return None


def find_header_row(sheet):
    """
    Search the first HEADER_SCAN_ROWS rows for a row that contains
    Part Number, Product Description, Data Rate, and Form Factor headers.
    """
    max_scan = min(sheet.max_row, HEADER_SCAN_ROWS)

    for row_number in range(1, max_scan + 1):
        headers = [
            sheet.cell(row=row_number, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]

        part_col = find_column(headers, PART_NUMBER_HEADERS)
        description_col = find_column(headers, PRODUCT_DESCRIPTION_HEADERS)
        rate_col = find_column(headers, DATA_RATE_HEADERS)
        form_col = find_column(headers, FORM_FACTOR_HEADERS)

        if part_col and description_col and rate_col and form_col:
            return {
                "row": row_number,
                "headers": headers,
                "part_col": part_col,
                "description_col": description_col,
                "rate_col": rate_col,
                "form_col": form_col,
            }

    return None


def find_catalog_sheet(workbook):
    """
    Automatically find the worksheet and header row that contain
    the four required catalog columns.
    """
    if SHEET_NAME:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f'Worksheet "{SHEET_NAME}" was not found. '
                f"Available sheets: {workbook.sheetnames}"
            )

        sheet = workbook[SHEET_NAME]
        header_info = find_header_row(sheet)

        if not header_info:
            raise ValueError(
                f'Could not find Part Number, Product Description, Data Rate, '
                f'and Form Factor headers in the first {HEADER_SCAN_ROWS} rows '
                f'of worksheet "{SHEET_NAME}".'
            )

        return sheet, header_info

    # Auto-detect across all tabs
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_info = find_header_row(sheet)

        if header_info:
            return sheet, header_info

    raise ValueError(
        "Could not find a worksheet containing all four required headers: "
        "Part Number, Product Description, Data Rate, and Form Factor. "
        f"Searched the first {HEADER_SCAN_ROWS} rows of these worksheets: "
        + ", ".join(workbook.sheetnames)
    )


def data_rate_value(rate):
    """Numeric value used for sorting largest to smallest."""
    text = clean_cell(rate).upper()

    if not text or text == "N/A":
        return -1

    matches = re.findall(r"(\d+(?:\.\d+)?)\s*(T|G|M|K)?", text)

    if not matches:
        return -1

    values = []

    for number_text, unit in matches:
        number = float(number_text)

        multiplier = {
            "T": 1_000_000,
            "G": 1_000,
            "M": 1,
            "K": 0.001,
            "": 1,
        }[unit]

        values.append(number * multiplier)

    return max(values)


def main():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Could not find {EXCEL_FILE}. "
            "Make sure catalog.xlsx is in the repository root."
        )

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
    )

    sheet, header_info = find_catalog_sheet(workbook)

    header_row = header_info["row"]
    part_number_column = header_info["part_col"]
    product_description_column = header_info["description_col"]
    data_rate_column = header_info["rate_col"]
    form_factor_column = header_info["form_col"]

    print(f'Catalog worksheet: "{sheet.title}"')
    print(f"Header row: {header_row}")
    print(
        "Columns: "
        f"Part Number={part_number_column}, "
        f"Product Description={product_description_column}, "
        f"Data Rate={data_rate_column}, "
        f"Form Factor={form_factor_column}"
    )

    catalog = []

    for row_number in range(header_row + 1, sheet.max_row + 1):
        part_number = clean_cell(
            sheet.cell(row=row_number, column=part_number_column).value
        )
        product_description = clean_cell(
            sheet.cell(row=row_number, column=product_description_column).value
        )
        data_rate = clean_cell(
            sheet.cell(row=row_number, column=data_rate_column).value
        )
        form_factor = clean_cell(
            sheet.cell(row=row_number, column=form_factor_column).value
        )

        # Customer catalog should not contain blank part numbers.
        if not part_number:
            continue

        catalog.append({
            "partNumber": part_number,
            "productDescription": product_description,
            "dataRate": data_rate,
            "formFactor": form_factor,
        })

    # Largest data rate first.
    # Stable sort preserves Excel order within the same data rate.
    catalog.sort(
        key=lambda item: data_rate_value(item["dataRate"]),
        reverse=True,
    )

    OUTPUT_FILE.write_text(
        json.dumps(catalog, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print(
        f"Generated {OUTPUT_FILE} with {len(catalog)} catalog items."
    )


if __name__ == "__main__":
    main()
